"""
Import script for raw_data.xlsx
Maps Excel columns to database schema and imports customers and loans
"""
import sys
import os
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal
import openpyxl
from sqlalchemy.orm import Session

# Set database URL to use port 5433
os.environ['DATABASE_URL'] = 'postgresql://ifrs9:ifrs9pass@localhost:5433/ifrs9'

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.db.session import SessionLocal
from src.db.models import Customer, FinancialInstrument, CustomerType, InstrumentType, InstrumentStatus, Stage


def excel_date_to_python(excel_date):
    """Convert Excel date (days since 1900-01-01) to Python date"""
    if excel_date is None or excel_date == '':
        return None
    
    if isinstance(excel_date, (date, datetime)):
        return excel_date if isinstance(excel_date, date) else excel_date.date()
    
    try:
        # Excel dates are days since 1900-01-01 (with leap year bug)
        excel_epoch = datetime(1899, 12, 30)
        return (excel_epoch + pd.Timedelta(days=int(excel_date))).date()
    except:
        return None


def map_customer_type(product_type: str) -> CustomerType:
    """Map product type to customer type"""
    if not product_type:
        return CustomerType.RETAIL
    
    product_type = product_type.upper()
    
    if 'SME' in product_type or 'SPECIAL PROGRAM' in product_type:
        return CustomerType.SME
    elif 'CORPORATE' in product_type:
        return CustomerType.CORPORATE
    else:
        return CustomerType.RETAIL


def map_instrument_type(product_desc: str) -> InstrumentType:
    """Map product description to instrument type"""
    if not product_desc:
        return InstrumentType.TERM_LOAN
    
    product_desc = product_desc.upper()
    
    if 'OVERDRAFT' in product_desc:
        return InstrumentType.OVERDRAFT
    elif 'BOND' in product_desc:
        return InstrumentType.BOND
    elif 'COMMITMENT' in product_desc:
        return InstrumentType.COMMITMENT
    else:
        return InstrumentType.TERM_LOAN


def map_status(status: str) -> InstrumentStatus:
    """Map loan status to instrument status"""
    if not status:
        return InstrumentStatus.ACTIVE
    
    status = status.upper()
    
    if status == 'CLOSED' or status == 'PAID OFF':
        return InstrumentStatus.DERECOGNIZED
    elif status == 'WRITTEN OFF':
        return InstrumentStatus.WRITTEN_OFF
    else:
        return InstrumentStatus.ACTIVE


def import_raw_data(excel_path: str, db: Session):
    """Import data from raw_data.xlsx"""
    
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Sheet1']
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"Found {len(headers)} columns")
    
    # Track unique customers
    customers_dict = {}
    instruments_list = []
    
    # Process rows
    row_count = 0
    skipped_count = 0
    
    print("\nProcessing rows...")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None for cell in row):
            continue
        
        row_count += 1
        
        # Create row dict
        row_data = dict(zip(headers, row))
        
        try:
            # Extract customer data
            customer_id = str(row_data.get('CUST_ID', '')).strip()
            if not customer_id or customer_id == 'None':
                print(f"Row {row_idx}: Skipping - no customer ID")
                skipped_count += 1
                continue
            
            # Add customer if not already tracked
            if customer_id not in customers_dict:
                customer = Customer(
                    customer_id=customer_id,
                    customer_name=str(row_data.get('ACCT_NM', 'Unknown')).strip(),
                    customer_type=map_customer_type(row_data.get('PRODUCT_TYPE')),
                    country='Uganda',
                    region=None,
                    industry_sector=None,
                    credit_rating=row_data.get('RISK_CLASSFICATION'),
                    is_watchlist=False,
                    is_defaulted=row_data.get('RISK_CLASSFICATION') == 'LOSS'
                )
                customers_dict[customer_id] = customer
            
            # Extract instrument data
            instrument_id = str(row_data.get('ACCT_NO', '')).strip()
            if not instrument_id or instrument_id == 'None':
                print(f"Row {row_idx}: Skipping - no account number")
                skipped_count += 1
                continue
            
            # Parse dates
            origination_date = row_data.get('START_DT')
            if isinstance(origination_date, (int, float)):
                from datetime import timedelta
                origination_date = date(1899, 12, 30) + timedelta(days=int(origination_date))
            elif not isinstance(origination_date, date):
                origination_date = date.today()
            
            maturity_date = row_data.get('MATURITY_DT')
            if isinstance(maturity_date, (int, float)):
                from datetime import timedelta
                maturity_date = date(1899, 12, 30) + timedelta(days=int(maturity_date))
            elif not isinstance(maturity_date, date):
                maturity_date = origination_date
            
            # Parse amounts
            principal_amount = row_data.get('PRINCIPALCLOSING') or row_data.get('PRINCIPAL_OPENING') or 0
            if principal_amount is None:
                principal_amount = 0
            principal_amount = Decimal(str(principal_amount))
            
            interest_rate = row_data.get('INTEREST_RATE') or 12.0
            if interest_rate is None:
                interest_rate = 12.0
            interest_rate = Decimal(str(interest_rate))
            
            # Parse DPD
            days_past_due = row_data.get('DAYS_IN_ARREAS') or 0
            if days_past_due is None:
                days_past_due = 0
            days_past_due = int(days_past_due)
            
            # Determine stage based on DPD
            if days_past_due > 90:
                current_stage = Stage.STAGE_3
            elif days_past_due > 30:
                current_stage = Stage.STAGE_2
            else:
                current_stage = Stage.STAGE_1
            
            # Create instrument
            instrument = FinancialInstrument(
                instrument_id=instrument_id,
                instrument_type=map_instrument_type(row_data.get('PROD_DESC')),
                customer_id=customer_id,
                origination_date=origination_date,
                maturity_date=maturity_date,
                principal_amount=principal_amount,
                interest_rate=interest_rate,
                currency=row_data.get('CRNCY_CD', 'UGX'),
                status=map_status(row_data.get('STATUS')),
                days_past_due=days_past_due,
                current_stage=current_stage,
                stage_date=date.today(),
                is_modified=row_data.get('IS_LOAN_RESTRUCTURED') == 'YES',
                is_poci=False
            )
            
            instruments_list.append(instrument)
            
            if row_count % 100 == 0:
                print(f"Processed {row_count} rows...")
        
        except Exception as e:
            print(f"Row {row_idx}: Error - {str(e)}")
            skipped_count += 1
            continue
    
    print(f"\nProcessed {row_count} rows")
    print(f"Skipped {skipped_count} rows")
    print(f"Found {len(customers_dict)} unique customers")
    print(f"Found {len(instruments_list)} instruments")
    
    # Insert into database
    print("\nInserting customers into database...")
    for customer in customers_dict.values():
        db.merge(customer)
    
    db.commit()
    print(f"Inserted {len(customers_dict)} customers")
    
    print("\nInserting instruments into database...")
    batch_size = 100
    for i in range(0, len(instruments_list), batch_size):
        batch = instruments_list[i:i+batch_size]
        for instrument in batch:
            db.merge(instrument)
        db.commit()
        print(f"Inserted {min(i+batch_size, len(instruments_list))}/{len(instruments_list)} instruments...")
    
    print("\n✅ Import completed successfully!")
    print(f"Total customers: {len(customers_dict)}")
    print(f"Total instruments: {len(instruments_list)}")


def main():
    """Main entry point"""
    excel_path = Path(__file__).parent.parent / 'test_data' / 'raw_data.xlsx'
    
    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    print("=" * 60)
    print("IFRS 9 Platform - Raw Data Import")
    print("=" * 60)
    
    # Create database session
    db = SessionLocal()
    
    try:
        import_raw_data(str(excel_path), db)
    except Exception as e:
        print(f"\n❌ Import failed: {str(e)}")
        import traceback
        traceback.print_exc()
        db.rollback()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
